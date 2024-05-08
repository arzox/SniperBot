'''
# Ethereum
refresh every minutes
Get all tokens created in the refresh time and their audits and overall information
If a token is promising add it to the sheet file
'''
import json
from datetime import *
from openpyxl import Workbook
from time import sleep


def load_api_keys(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)


def security_check(api, token, chain, debug=False):
    audit = api.get_token_audit(chain, token["address"]).get("data")
    sleep(0.5)
    info = api.get_token_info(chain, token["address"]).get("data")
    try:
        # Checks on audit data
        if audit.get("isPotentiallyScam") == "yes":
            raise ValueError(f"The value isPotentiallyScam = {audit.get('isPotentiallyScam')}")
        if audit.get("isHoneypot") == "yes":
            raise ValueError(f"The value isHoneypot = {audit.get('isHoneypot')}")
        if audit.get("isContractRenounced") == "no" or audit.get("isContractRenounced") == "warning" or audit.get(
                "isContractRenounced") is None:
            raise ValueError(f"The value isContractRenounced = {audit.get('isContractRenounced')}")
        if audit.get("sellTax").get("max") > 0.02 or audit.get("sellTax").get("max") > 0.02:
            raise ValueError(f"The tax value : Sell = {audit.get('sellTax')}  -  Buy = {audit.get('buyTax')}")
        # Checks on info data
        if info.get("holders", 0) < 10:
            raise ValueError(f"The value holders = {info.get('holders')}")

        return {"address": token["address"], **audit, **info}
    except ValueError as e:
        if debug:
            print(f"Security check failed: {e} - Check audit data.")
        return False
    except Exception as e:
        if debug:
            print(f"Security check failed: {e} - Check info data.")
        return False


def link_last_cell(ws, column, address=None, chain="ether"):
    address_cell = ws.cell(row=ws.max_row, column=column)
    address_cell.hyperlink = "https://www.dextools.io/app/en/{}/pair-explorer/{}".format(chain, address)
    address_cell.value = "dexTools link"
    address_cell.style = "Hyperlink"


def store_token(info, token, chain, worksheet):
    worksheet.append(
        [token.get("symbol"), info["address"],
         "https://www.dextools.io/app/en/{}/pair-explorer/{}".format(chain, info["address"]),
         info["mcap"], info["fdv"], info["holders"]])
    link_last_cell(worksheet, 3, info["address"])
    worksheet.parent.save("tokens.xlsx")


def get_token_list(api, chain, time_range, number=50):
    now = datetime.now(timezone.utc)
    tokens_list = api.get_tokens(chain, from_=(now - time_range).isoformat(),
                                 to=now.isoformat(),
                                 pageSize=number, sort="creationTime", order="desc")
    sleep(0.5)
    print(f"{len(tokens_list.get('data').get('tokens'))} tokens found")
    return tokens_list


def set_sheet():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Tokens"

    ws.append(["Symbol", "Address", "Link", "mcap", "fdv", "Holders"])
    return workbook
